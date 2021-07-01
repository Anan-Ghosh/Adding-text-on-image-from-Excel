from openpyxl import load_workbook
from PIL import Image, ImageFont, ImageDraw

file = load_workbook("MidNightHackathon212.1.xlsx")

file_obj = file.active

max_col = file_obj.max_column
max_row = file_obj.max_row

title_font = ImageFont.truetype('Yellowtail/Yellowtail-Regular.ttf',size=80)



list=[]

dict={}


for i in range(2, max_row+1):
    name = file_obj.cell(row=i, column=1)
    id = file_obj.cell(row=i, column=2)
    dict[name.value] = id.value

for x in dict:
    my_image = Image.open("ben-sweet-2LowviVHZ-E-unsplash-1.jpeg")
    image_editable = ImageDraw.Draw(my_image)
    title_text = "name:"+x
    title_text1 = "id:" + dict[x]
    image_editable.text((15, 15), title_text, (237, 230, 211), font=title_font)
    image_editable.text((15, 70), title_text1, (237, 230, 211), font=title_font)
    my_image.save(dict[x]+".jpg")











